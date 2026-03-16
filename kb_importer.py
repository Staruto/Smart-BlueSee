"""
KB Importer for private sources (email, Moodle exports, misc docs).

Converts supported inputs into [Section]-formatted .txt files under
knowledge/import/* so existing KnowledgeBase loader can index them.

Supported now:
  - Email: .mbox, .eml
  - Moodle exports: .html, .htm
  - Docs: .txt, .md, .csv

Optional (if dependencies are installed):
  - .pdf via pypdf
  - .docx via python-docx
  - .xlsx via openpyxl
"""

from __future__ import annotations

import argparse
import datetime as dt
import email
import hashlib
import json
import mailbox
import os
import re
from dataclasses import dataclass
from email import policy
from email.parser import BytesParser
from html import unescape
from typing import Dict, Iterable, List, Optional, Tuple

from config import (
	ENABLE_IMPORT_DEDUP,
	ENABLE_IMPORT_PII_REDACTION,
	IMPORT_DOCS_OUTPUT_DIR,
	IMPORT_EMAIL_OUTPUT_DIR,
	IMPORT_MANIFEST_DIR,
	IMPORT_MAX_ITEMS_PER_RUN,
	IMPORT_MIN_CONTENT_CHARS,
	IMPORT_MOODLE_OUTPUT_DIR,
)

try:
	from bs4 import BeautifulSoup  # type: ignore
except Exception:
	BeautifulSoup = None


PII_PATTERNS: Dict[str, re.Pattern[str]] = {
	"email": re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b"),
	"student_id": re.compile(r"\b\d{7,10}\b"),
	"phone": re.compile(r"\b(?:\+?\d[\d\- ]{8,}\d)\b"),
	"token": re.compile(r"(?i)\b(token|api[_-]?key|secret|password)\b\s*[:=]\s*\S+"),
}

FORWARDED_LINE_RE = re.compile(
	r"^(from|sent|to|cc|subject)\s*:\s*", re.IGNORECASE | re.MULTILINE
)
TAG_RE = re.compile(r"<[^>]+>")
SPACE_RE = re.compile(r"\s+")


@dataclass
class ImportRecord:
	source_type: str
	source_path: str
	output_path: str
	title: str
	chars: int
	redaction_flags: List[str]
	duplicate: bool
	status: str
	error: str = ""


def _ensure_dir(path: str) -> None:
	os.makedirs(path, exist_ok=True)


def _safe_name(text: str) -> str:
	clean = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_")
	return clean[:90] or "item"


def _now() -> str:
	return dt.datetime.now().strftime("%Y-%m-%d %H:%M")


def _normalize_text(text: str) -> str:
	text = text.replace("\r\n", "\n").replace("\r", "\n")
	text = unescape(text)
	text = re.sub(r"\n{3,}", "\n\n", text)
	text = SPACE_RE.sub(" ", text)
	return text.strip()


def _clean_email_text(text: str) -> str:
	text = FORWARDED_LINE_RE.sub("", text)
	return _normalize_text(text)


def _redact(text: str, enabled: bool) -> Tuple[str, List[str]]:
	if not enabled:
		return text, []
	flags: List[str] = []
	out = text
	for label, pattern in PII_PATTERNS.items():
		if pattern.search(out):
			flags.append(label)
			out = pattern.sub(f"[REDACTED_{label.upper()}]", out)
	return out, flags


def _content_hash(text: str) -> str:
	return hashlib.sha256(_normalize_text(text).lower().encode("utf-8")).hexdigest()


def _format_section(title: str, source: str, body: str) -> str:
	return (
		f"[{title}]\n"
		f"Source: {source}\n"
		f"Last updated: {_now()}\n\n"
		f"{body}\n"
	)


def _extract_html_text(raw_html: str) -> str:
	if BeautifulSoup is not None:
		soup = BeautifulSoup(raw_html, "html.parser")
		for tag in soup(["script", "style", "noscript"]):
			tag.decompose()
		text = soup.get_text(separator="\n", strip=True)
		return _normalize_text(text)
	text = TAG_RE.sub(" ", raw_html)
	return _normalize_text(text)


def _write_section(
	output_dir: str,
	filename_stem: str,
	title: str,
	source: str,
	body: str,
	dry_run: bool,
) -> str:
	_ensure_dir(output_dir)
	output_path = os.path.join(output_dir, f"{_safe_name(filename_stem)}.txt")
	if dry_run:
		return output_path
	with open(output_path, "w", encoding="utf-8") as f:
		f.write(_format_section(title, source, body))
	return output_path


def _email_body_from_message(msg: email.message.EmailMessage) -> str:
	if msg.is_multipart():
		parts = []
		for part in msg.walk():
			ctype = part.get_content_type()
			if ctype == "text/plain":
				payload = part.get_payload(decode=True)
				if payload:
					charset = part.get_content_charset() or "utf-8"
					parts.append(payload.decode(charset, errors="ignore"))
		if parts:
			return "\n\n".join(parts)
	payload = msg.get_payload(decode=True)
	if isinstance(payload, bytes):
		charset = msg.get_content_charset() or "utf-8"
		return payload.decode(charset, errors="ignore")
	if isinstance(payload, str):
		return payload
	return ""


def _iter_email_items(path: str) -> Iterable[Tuple[str, str, str, str]]:
	"""Yield (source_path, subject, date, body)."""
	if os.path.isfile(path) and path.lower().endswith(".mbox"):
		box = mailbox.mbox(path)
		for i, msg in enumerate(box):
			subject = msg.get("subject", "No Subject")
			date = msg.get("date", "")
			body = _email_body_from_message(msg)
			yield (f"{path}#{i}", subject, date, body)
		return

	if os.path.isfile(path) and path.lower().endswith(".eml"):
		with open(path, "rb") as f:
			msg = BytesParser(policy=policy.default).parse(f)
		subject = msg.get("subject", "No Subject")
		date = msg.get("date", "")
		body = _email_body_from_message(msg)
		yield (path, subject, date, body)
		return

	if os.path.isdir(path):
		for root, _, files in os.walk(path):
			for name in files:
				full = os.path.join(root, name)
				low = name.lower()
				if low.endswith(".eml"):
					yield from _iter_email_items(full)
				elif low.endswith(".mbox"):
					yield from _iter_email_items(full)


def import_email(
	source_path: str,
	dry_run: bool,
	redact: bool,
	dedup: bool,
	max_items: int,
) -> List[ImportRecord]:
	records: List[ImportRecord] = []
	seen: set[str] = set()

	for idx, (src, subject, date, body) in enumerate(_iter_email_items(source_path), 1):
		if max_items and idx > max_items:
			break
		title = f"Email: {subject.strip() or 'No Subject'}"
		cleaned = _clean_email_text(body)
		cleaned, flags = _redact(cleaned, enabled=redact)

		if len(cleaned) < IMPORT_MIN_CONTENT_CHARS:
			records.append(
				ImportRecord("email", src, "", title, len(cleaned), flags, False, "skipped_short")
			)
			continue

		h = _content_hash(cleaned)
		if dedup and h in seen:
			records.append(ImportRecord("email", src, "", title, len(cleaned), flags, True, "skipped_duplicate"))
			continue
		seen.add(h)

		stem = f"email_{idx}_{subject}_{date}"[:140]
		output = _write_section(
			IMPORT_EMAIL_OUTPUT_DIR,
			stem,
			title,
			f"{src} | date={date}",
			cleaned,
			dry_run,
		)
		records.append(ImportRecord("email", src, output, title, len(cleaned), flags, False, "imported"))

	return records


def _iter_moodle_html(path: str) -> Iterable[str]:
	if os.path.isfile(path) and path.lower().endswith((".html", ".htm")):
		yield path
		return
	if os.path.isdir(path):
		for root, _, files in os.walk(path):
			for name in files:
				if name.lower().endswith((".html", ".htm")):
					yield os.path.join(root, name)


def import_moodle(
	source_path: str,
	dry_run: bool,
	redact: bool,
	dedup: bool,
	max_items: int,
) -> List[ImportRecord]:
	records: List[ImportRecord] = []
	seen: set[str] = set()

	for idx, html_path in enumerate(_iter_moodle_html(source_path), 1):
		if max_items and idx > max_items:
			break
		try:
			with open(html_path, "r", encoding="utf-8", errors="ignore") as f:
				raw = f.read()
			text = _extract_html_text(raw)
			text, flags = _redact(text, enabled=redact)
			if len(text) < IMPORT_MIN_CONTENT_CHARS:
				records.append(
					ImportRecord("moodle", html_path, "", os.path.basename(html_path), len(text), flags, False, "skipped_short")
				)
				continue

			h = _content_hash(text)
			if dedup and h in seen:
				records.append(ImportRecord("moodle", html_path, "", os.path.basename(html_path), len(text), flags, True, "skipped_duplicate"))
				continue
			seen.add(h)

			title = f"Moodle: {os.path.splitext(os.path.basename(html_path))[0]}"
			output = _write_section(
				IMPORT_MOODLE_OUTPUT_DIR,
				f"moodle_{idx}_{os.path.basename(html_path)}",
				title,
				html_path,
				text,
				dry_run,
			)
			records.append(ImportRecord("moodle", html_path, output, title, len(text), flags, False, "imported"))
		except Exception as e:
			records.append(ImportRecord("moodle", html_path, "", os.path.basename(html_path), 0, [], False, "failed", str(e)))
	return records


def _extract_doc_text(path: str) -> str:
	ext = os.path.splitext(path)[1].lower()
	if ext in (".txt", ".md", ".csv"):
		with open(path, "r", encoding="utf-8", errors="ignore") as f:
			return _normalize_text(f.read())
	if ext in (".html", ".htm"):
		with open(path, "r", encoding="utf-8", errors="ignore") as f:
			return _extract_html_text(f.read())
	if ext == ".pdf":
		try:
			from pypdf import PdfReader  # type: ignore
		except Exception:
			return ""
		reader = PdfReader(path)
		return _normalize_text("\n".join((page.extract_text() or "") for page in reader.pages))
	if ext == ".docx":
		try:
			from docx import Document  # type: ignore
		except Exception:
			return ""
		doc = Document(path)
		return _normalize_text("\n".join(p.text for p in doc.paragraphs))
	if ext == ".xlsx":
		try:
			from openpyxl import load_workbook  # type: ignore
		except Exception:
			return ""
		wb = load_workbook(path, read_only=True, data_only=True)
		lines: List[str] = []
		for ws in wb.worksheets:
			lines.append(f"Sheet: {ws.title}")
			for row in ws.iter_rows(values_only=True):
				vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
				if vals:
					lines.append(" | ".join(vals))
		return _normalize_text("\n".join(lines))
	return ""


def import_docs(
	source_path: str,
	dry_run: bool,
	redact: bool,
	dedup: bool,
	max_items: int,
) -> List[ImportRecord]:
	records: List[ImportRecord] = []
	seen: set[str] = set()
	allowed_ext = {".txt", ".md", ".csv", ".pdf", ".docx", ".xlsx", ".html", ".htm"}
	files: List[str] = []

	if os.path.isfile(source_path):
		files = [source_path]
	elif os.path.isdir(source_path):
		for root, _, names in os.walk(source_path):
			for n in names:
				if os.path.splitext(n)[1].lower() in allowed_ext:
					files.append(os.path.join(root, n))

	for idx, doc_path in enumerate(sorted(files), 1):
		if max_items and idx > max_items:
			break
		try:
			text = _extract_doc_text(doc_path)
			text, flags = _redact(text, enabled=redact)
			if len(text) < IMPORT_MIN_CONTENT_CHARS:
				records.append(ImportRecord("docs", doc_path, "", os.path.basename(doc_path), len(text), flags, False, "skipped_short_or_unsupported"))
				continue

			h = _content_hash(text)
			if dedup and h in seen:
				records.append(ImportRecord("docs", doc_path, "", os.path.basename(doc_path), len(text), flags, True, "skipped_duplicate"))
				continue
			seen.add(h)

			title = f"Doc: {os.path.splitext(os.path.basename(doc_path))[0]}"
			output = _write_section(
				IMPORT_DOCS_OUTPUT_DIR,
				f"doc_{idx}_{os.path.basename(doc_path)}",
				title,
				doc_path,
				text,
				dry_run,
			)
			records.append(ImportRecord("docs", doc_path, output, title, len(text), flags, False, "imported"))
		except Exception as e:
			records.append(ImportRecord("docs", doc_path, "", os.path.basename(doc_path), 0, [], False, "failed", str(e)))

	return records


def _save_manifest(records: List[ImportRecord], dry_run: bool) -> str:
	batch_id = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
	payload = {
		"batch_id": batch_id,
		"generated_at": dt.datetime.now().isoformat(),
		"dry_run": dry_run,
		"summary": {
			"total": len(records),
			"imported": sum(1 for r in records if r.status == "imported"),
			"skipped": sum(1 for r in records if r.status.startswith("skipped")),
			"failed": sum(1 for r in records if r.status == "failed"),
		},
		"records": [r.__dict__ for r in records],
	}

	_ensure_dir(IMPORT_MANIFEST_DIR)
	path = os.path.join(IMPORT_MANIFEST_DIR, f"import_manifest_{batch_id}.json")
	if not dry_run:
		with open(path, "w", encoding="utf-8") as f:
			json.dump(payload, f, ensure_ascii=False, indent=2)
	return path


def main() -> None:
	parser = argparse.ArgumentParser(
		description="Import email/Moodle/docs into Section-formatted knowledge files"
	)
	parser.add_argument("--email-path", type=str, help="Path to .mbox/.eml file or directory")
	parser.add_argument("--moodle-path", type=str, help="Path to Moodle html export file or directory")
	parser.add_argument("--docs-path", type=str, help="Path to document file or directory")
	parser.add_argument("--dry-run", action="store_true", help="Parse and report only, do not write files")
	parser.add_argument("--no-redact", action="store_true", help="Disable PII redaction (not recommended)")
	parser.add_argument("--no-dedup", action="store_true", help="Disable duplicate filtering")
	parser.add_argument("--max-items", type=int, default=IMPORT_MAX_ITEMS_PER_RUN, help="Limit processed items per source (0=all)")
	args = parser.parse_args()

	redact = ENABLE_IMPORT_PII_REDACTION and not args.no_redact
	dedup = ENABLE_IMPORT_DEDUP and not args.no_dedup
	max_items = max(0, args.max_items)

	records: List[ImportRecord] = []

	if args.email_path:
		records.extend(import_email(args.email_path, args.dry_run, redact, dedup, max_items))
	if args.moodle_path:
		records.extend(import_moodle(args.moodle_path, args.dry_run, redact, dedup, max_items))
	if args.docs_path:
		records.extend(import_docs(args.docs_path, args.dry_run, redact, dedup, max_items))

	if not (args.email_path or args.moodle_path or args.docs_path):
		parser.error("At least one source path is required (--email-path/--moodle-path/--docs-path)")

	manifest_path = _save_manifest(records, dry_run=args.dry_run)

	imported = sum(1 for r in records if r.status == "imported")
	skipped = sum(1 for r in records if r.status.startswith("skipped"))
	failed = sum(1 for r in records if r.status == "failed")
	print("=" * 60)
	print("KB Importer Summary")
	print(f"  Total   : {len(records)}")
	print(f"  Imported: {imported}")
	print(f"  Skipped : {skipped}")
	print(f"  Failed  : {failed}")
	if args.dry_run:
		print("  Manifest: dry-run mode (not written)")
	else:
		print(f"  Manifest: {manifest_path}")
	print("=" * 60)


if __name__ == "__main__":
	main()
